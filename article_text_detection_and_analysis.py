from newspaper import Article, ArticleException
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font
from nltk import *
from functions import *
from pathlib import Path
import os


Output_Data = []

# Step 1: EXTRACTING AND STORING THE STOPWORDS AND MASTER DICTIONARY KEY WORDS IN SET

sw_paths = [
    "Resources/StopWords/StopWords_Auditor.txt",
    "Resources/StopWords/StopWords_Currencies.txt",
    "Resources/StopWords/StopWords_DatesandNumbers.txt",
    "Resources/StopWords/StopWords_Generic.txt",
    "Resources/StopWords/StopWords_GenericLong.txt",
    "Resources/StopWords/StopWords_Geographic.txt",
    "Resources/StopWords/StopWords_Names.txt"
]

pos_md_path = "Resources/MasterDictionary/positive-words.txt"
neg_md_path = "Resources/MasterDictionary/negative-words.txt"

stopwords_cstm = set()

punctuation = ['.', ',', ';', ':', "-", "?"]
stopwords_cstm.update(punctuation)

for path in sw_paths:
    with open(path, 'r') as file:
        stopwords_cstm.update(map(str.lower, file.read().split()))


pos_dict = load_master_dict(pos_md_path)
neg_dict = load_master_dict(neg_md_path)

directory_name = "StoredDataFiles"

if not os.path.exists(directory_name):
    os.makedirs(directory_name)
    print(f"Directory '{directory_name}' created successfully.")
else:
    print(f"Directory '{directory_name}' already exists.")

# Step 2: EXTRACTING THE EXCEL DATA AND STORING THE IDS AND URLS IN A LIST

if is_directory_empty(directory_name):
    file_path = "Resources/input.xlsx"
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    for row in sheet.iter_rows(min_row=2,values_only=True):

        output_data_row = []
        id = row[0]
        url = row[1]
        output_data_row.append(id)
        output_data_row.append(url)
        article_title, article_text = getArticleData(url)
        saveArticleData(id, article_title, article_text)
        if article_text != None and article_text != None:
            filepath = f'StoredDataFiles/{id}.txt'
            with open(filepath, 'r', encoding='utf-8') as file:
                text = file.read()
            (   
            positive_score,
            negative_score,
            polarity_score,
            subjectivity_score,
            Average_Number_of_Words_Per_Sentence,
            percentage_of_Complex_words,
            Fog_Index,
            Complex_Word_Count,
            Word_Count,
            Syllable_Count_Per_Word,
            Personal_Pronouns,
            Average_Word_Length
            ) = main_function(text, stopwords_cstm, pos_dict, neg_dict)

            output_data_row.append(positive_score)
            output_data_row.append(negative_score)
            output_data_row.append(polarity_score)
            output_data_row.append(subjectivity_score)
            output_data_row.append(Average_Number_of_Words_Per_Sentence)
            output_data_row.append(percentage_of_Complex_words)
            output_data_row.append(Fog_Index)
            output_data_row.append(Average_Number_of_Words_Per_Sentence)
            output_data_row.append(Complex_Word_Count)
            output_data_row.append(Word_Count)
            output_data_row.append(Syllable_Count_Per_Word)
            output_data_row.append(Personal_Pronouns)
            output_data_row.append(Average_Word_Length)
        else:
            for i in range(13):
                output_data_row.append("NA")


            

        Output_Data.append(output_data_row)

        
    workbook.close()


# STEP 3: ANALYSING THE SENTIMMENTS WORDS

wb = Workbook()
sheet = wb.create_sheet(title="Output Data Structure")
headings = [
    "URL_ID", "URL", "POSITIVE SCORE", "NEGATIVE SCORE", "POLARITY SCORE", "SUBJECTIVITY SCORE",
    "AVG SENTENCE LENGTH", "PERCENTAGE OF COMPLEX WORDS", "FOG INDEX", "AVG NUMBER OF WORDS PER SENTENCE",
    "COMPLEX WORD COUNT", "WORD COUNT", "SYLLABLE PER WORD", "PERSONAL PRONOUNS", "AVG WORD LENGTH"
]
sheet.append(headings)
for cell in sheet[1]:
    cell.font = Font(bold=True)

for row in Output_Data:
    sheet.append(row)

for cell in sheet.iter_cols(min_row = 2,min_col=2, max_col = 2, values_only=False):
    for cell_in_column in cell:
        url = cell_in_column.value 
        cell_in_column.font = Font(underline="single", color="0563C1")
        cell_in_column.hyperlink = url
output_file = "Output Data Structure.xlsx"
wb.save(output_file)

wb.close()

store_path = "Output_Data.txt"
with open(store_path,"w") as file:
    for line in Output_Data:
        file.write(str(line) + "\n")







# print("Positive Score:                          ", positive_score)
# print("Negative Score:                          ", negative_score)
# print("Polarity Score:                          ", polarity_score)
# print("Subjectivity Score:                      ", subjectivity_score)

# print("Average Number of Words Per Sentence:    ", Average_Number_of_Words_Per_Sentence)
# print("Percentage of Complex Words:             ", percentage_of_Complex_words,"%")
# print("Fog Index:                               ", Fog_Index)

# print("Average Number of Words Per Sentence:    ", Average_Number_of_Words_Per_Sentence)

# print("Complex Word Count:                      ", Complex_Word_Count)

# print("Word count after cleaning:               ", Word_Count)

# print("Syllable Count Per Word:                 ", Syllable_Count_Per_Word)

# print("Personal Pronouns:                       ", Personal_Pronouns)

# print("Average Word Length:                     ", Average_Word_Length)







    











